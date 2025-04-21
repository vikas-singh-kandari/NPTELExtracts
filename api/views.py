from django.shortcuts import render, redirect
from django.http import FileResponse, JsonResponse
import os
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from pdf2image import convert_from_path
from .utils import extract_certificate_data
from django.contrib.auth.models import User
from .models import Certificate
from rest_framework import viewsets
from .serializer import CertificateSerializer
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages

def save_file(file, folder):
    """ Save uploaded file and return the file path """
    file_path = os.path.join(folder, file.name)
    with open(file_path, 'wb') as f:
        for chunk in file.chunks():
            f.write(chunk)
    return file_path


def convert_pdf_to_images(pdf_path, output_folder):
    """ Convert PDF to images and return image paths """
    images = convert_from_path(pdf_path, dpi=300)
    image_paths = []

    for i, img in enumerate(images):
        image_filename = f"{os.path.splitext(os.path.basename(pdf_path))[0]}_page_{i + 1}.png"
        image_path = os.path.join(output_folder, image_filename)
        img.save(image_path, 'PNG')
        image_paths.append(image_path)

    return image_paths


def initialize_excel(excel_file):
    """ Initialize Excel with headers and formatting """
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'NPTEL Result'

        # Matching the header sequence
        headers = [
            'S.No', 'Dept', 'Study Year', 'College Roll No', 'Student Name', 'Course Code', 'Course Name', 'Sem',
            'Year of passing', 'Score from Assignment (25)', 'Proctored Exam Score (75)', 'Final Exam Score (100)',
            'Eligible for modified pass', 'Week', 'Course Type Credit Transfer/Value Added',
            'Credits as per MRIIRS Policy', 'Grade (numeric) as per MRIIRS', 'Grade (letter) as per MRIIRS',
            'Timeline', 'Link for Certificate'
        ]
        
        # Add headers
        sheet.append(headers)

        # Apply bold font and center alignment
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        workbook.save(excel_file)


def format_excel(sheet):
    """ Auto-fit columns and apply header formatting """
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Set the column width with a buffer
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width


def upload_certificates(request):
    """ Upload, process, extract, and save certificates to Excel and DB """
    if not request.user.is_authenticated:
        return redirect('login')

    username = request.user.username
    excel_output_folder = os.path.join(settings.MEDIA_ROOT, 'excel_output')
    user_folder = os.path.join(excel_output_folder, username)
    os.makedirs(user_folder, exist_ok=True)

    excel_file = os.path.join(user_folder, f'NPTEL_{username}.xlsx')
    initialize_excel(excel_file)

    # Load Excel
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    extracted_data = []

    if request.method == 'POST':
        files = request.FILES.getlist('certificates')

        # Folder paths
        cert_images_folder = os.path.join(settings.MEDIA_ROOT, 'certificates_images')
        os.makedirs(cert_images_folder, exist_ok=True)

        for file in files:
            file_ext = os.path.splitext(file.name)[-1].lower()

            # Handle PDF or image files
            if file_ext == '.pdf':
                pdf_path = save_file(file, cert_images_folder)
                image_paths = convert_pdf_to_images(pdf_path, cert_images_folder)
            else:
                image_path = save_file(file, cert_images_folder)
                image_paths = [image_path]

            for img_path in image_paths:
                certificate_data = extract_certificate_data(img_path)

                # Generate image URL
                image_url = request.build_absolute_uri(
                    os.path.join(settings.MEDIA_URL, 'certificates_images', os.path.basename(img_path))
                )

                # Save data to the database
                Certificate.objects.create(
                    user=request.user,
                    dept=certificate_data.get('Dept', ''),
                    study_year=certificate_data.get('Study Year', ''),
                    college_roll_no=certificate_data.get('College Roll No', ''),
                    student_name=certificate_data.get('Student Name', ''),
                    course_code=certificate_data.get('Course Code', ''),
                    course_name=certificate_data.get('Course Name', ''),
                    sem=certificate_data.get('Sem', ''),
                    year_of_passing=certificate_data.get('Year of passing', ''),
                    assignment_score=certificate_data.get('Score from Assignment(25)', ''),
                    proctored_exam_score=certificate_data.get('Proctored Exam Score(75)', ''),
                    final_exam_score=certificate_data.get('Final Exam Score(100)', ''),
                    eligible_for_modified_pass=certificate_data.get('Eligible for modified pass', ''),
                    week=certificate_data.get('Week', ''),
                    credit_transfer=certificate_data.get('Course Type Credit Transfer/Value Added', ''),
                    credits_mriirs=certificate_data.get('Credits as per MRIIRS Policy', ''),
                    grade_numeric=certificate_data.get('Grade(numeric) as per MRIIRS', ''),
                    grade_letter=certificate_data.get('Grade(letter) as per MRIIRS', ''),
                    timeline=certificate_data.get('Timeline', ''),
                    image_url=image_url
                )

                # Insert data into Excel
                row_data = [
                    next_row - 1,
                    certificate_data.get('Dept', ''),
                    certificate_data.get('Study Year', ''),
                    certificate_data.get('College Roll No', ''),
                    certificate_data.get('Student Name', ''),
                    certificate_data.get('Course Code', ''),
                    certificate_data.get('Course Name', ''),
                    certificate_data.get('Sem', ''),
                    certificate_data.get('Year of passing', ''),
                    certificate_data.get('Score from Assignment(25)', ''),
                    certificate_data.get('Proctored Exam Score(75)', ''),
                    certificate_data.get('Final Exam Score(100)', ''),
                    certificate_data.get('Eligible for modified pass', ''),
                    certificate_data.get('Week', ''),
                    certificate_data.get('Course Type Credit Transfer/Value Added', ''),
                    certificate_data.get('Credits as per MRIIRS Policy', ''),
                    certificate_data.get('Grade(numeric) as per MRIIRS', ''),
                    certificate_data.get('Grade(letter) as per MRIIRS', ''),
                    certificate_data.get('Timeline', ''),
                    image_url
                ]

                # Append data to Excel
                sheet.append(row_data)

                # Add data for frontend display
                extracted_data.append({
                    'S_NO': next_row - 1,
                    'Student_Name': certificate_data.get('Student Name', ''),
                    'Course_Code': certificate_data.get('Course Code', ''),
                    'Course_Name': certificate_data.get('Course Name', ''),
                    'Assignment_Score': certificate_data.get('Score from Assignment(25)', ''),
                    'Proctored_Score': certificate_data.get('Proctored Exam Score(75)', ''),
                    'Final_Exam_Score': certificate_data.get('Final Exam Score(100)', ''),
                    'Image_URL': image_url
                })

                next_row += 1

        # Auto-fit column widths and save Excel
        format_excel(sheet)
        workbook.save(excel_file)

        return render(request, 'home.html', {'data': extracted_data, 'message': 'Certificates uploaded successfully!'})

    return render(request, 'home.html')


def download_excel(request):
    """ Download Excel file """
    if not request.user.is_authenticated:
        return redirect('login')

    username = request.user.username
    excel_output_folder = os.path.join(settings.MEDIA_ROOT, 'excel_output')
    user_folder = os.path.join(excel_output_folder, username)
    excel_file = os.path.join(user_folder, f'NPTEL_{username}.xlsx')

    if os.path.exists(excel_file):
        return FileResponse(open(excel_file, 'rb'), as_attachment=True, filename=f'NPTEL_{username}.xlsx')
    else:
        return JsonResponse({'error': 'Excel file not found'}, status=404)



class CertificateViewSet(viewsets.ModelViewSet):
    queryset = Certificate.objects.all()
    serializer_class = CertificateSerializer



# =================================Login User===================


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('/')  
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'login.html')

def signup_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            login(request, user)
            return redirect('/') 
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = UserCreationForm()
    
    return render(request, 'signup.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')


